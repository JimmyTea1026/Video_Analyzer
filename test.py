from analyzer.analyzer import Analyzer
from analyzer.face_detector.face_detector import FaceDetector
from mask_test.StabilizationAlgorithm import StabilizationAlgorithm
from openpyxl import Workbook
import os
import pickle
import cv2
import matplotlib.pyplot as plt
import numpy as np

def mask_detect(analyzer, frame, rect):
    x1, y1, x2, y2 = rect
    h = y2 - y1
    w = x2 - x1
    if max(h, w) < 100:
        return None
    model_result = analyzer.inference_mask(frame)
    return model_result

def face_detect(frame):
    face_detector = FaceDetector('scrfd')
    rect = face_detector.detect(frame)
    return rect

def image_combination(frame, rect, model_result, algo_result):
    model_image = draw_image(frame, rect, model_result)
    algo_image = draw_image(frame, rect, algo_result)
    combined_image = cv2.vconcat([model_image, algo_image])
    cv2.putText(combined_image, "Mask Detection", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
    return combined_image

def draw_image(frame, rect, result):
    if rect is None:
        return frame
    draw_frame = frame.copy()
    x1, y1, x2, y2 = rect
    color = (0, 0, 0) 
    if result is None: color = (255, 255, 255)
    else: color = (0, 255, 0) if result else (0, 0, 255)
    
    cv2.rectangle(draw_frame, (x1, y1), (x2, y2), color, 2)
    if result is None:
        cv2.putText(draw_frame, "No Face", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)
    else:
        cv2.putText(draw_frame, "Mask" if result else "No Mask", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)
    return draw_frame

def video_generate(video_name, combined_image_list):
    path = "./mask_test/video"
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    width, height = combined_image_list[0].shape[1], combined_image_list[0].shape[0]
    fps = 15.0
    out = cv2.VideoWriter(f'{path}/{video_name}_result.mp4', fourcc, fps, (width, height))
    for image in combined_image_list:
        out.write(image)
    out.release()

def chart_plot(model_results, algo_result_list, video):
    def transform(results):
        data = []
        for result in results:
            if result is None:
                data.append(0)
            else:
                data.append(1 if result else -1)
        return data
    
    model_data = transform(model_results)
    fig, axes = plt.subplots(2,3)
    axs = axes.flat
    axs[0].plot(model_data, label="Model", color="black", linewidth=0.5)
    for i, algo_results in enumerate(algo_result_list):
        ax = axs[i+1]
        algo_data = transform(algo_results)
        ax.plot(algo_data, label="Algorithm", color="red", linewidth=0.5)
        ax.set_title(f"Queue Size {i+10}")
    plt.tight_layout()
    plt.savefig(f"./mask_test/plot/{video}.png")

def queue_size_test(model_results, ws, idx, video):
    result_size = sum(1 for result in model_results if result is not None)
    ws.cell(row=1, column=1, value="Result Size")
    ws.cell(row=1, column=2, value="Model Accuracy")
    start = 10
    end = 15
    titles = ["Queue Size {}".format(i) for i in range(start, end)]
    for i, title in enumerate(titles, start=3):  # 從第三列開始
        ws.cell(row=1, column=i, value=title)
        
    model_acc = sum(1 for result in model_results if result == True) / result_size
    model_acc = round(model_acc, 3)
    print(f"Model accuracy: {model_acc}")
    ws.cell(row=idx+2, column=1, value=result_size)
    ws.cell(row=idx+2, column=2, value=model_acc)
    algo_result_list = []
    column_count = 3
    for i in range(start, end):
        algorithm = StabilizationAlgorithm(q_size=i)
        algo_results = []
        for result in model_results:
            algo_results.append(algorithm.get_algorithm_result(result))
        algo_acc = sum(1 for _ in iter(algo_results) if _ == True) / result_size
        algo_acc = round(algo_acc, 3)
        # print(f"Queue size: {i} Algorithm accuracy: {algo_acc}")
        ws.cell(row=idx+2, column=column_count, value=algo_acc)
        algo_result_list.append(algo_results)
        column_count += 1
    chart_plot(model_results, algo_result_list, video)
        
def queue_test():
    wb = Workbook()
    ws = wb.active
    video_dir_path = "test/Mask/mask"
    videos = os.listdir(video_dir_path)
    model_result_list = pickle.load(open("/mask_test/model_result_list.pkl", "rb"))
    for idx, video in enumerate(videos):
        print(f"Processing {video}")
        model_results = model_result_list[video]
        queue_size_test(model_results, ws, idx, video)
    wb.save("/mask_test/mask_detection_result.xlsx")
            
def cal_model_result(video_dir_path):
    analyzer = Analyzer()
    model_result_list = {}
    rect_list = {}
    videos = os.listdir(video_dir_path)
    for video in videos:
        video_path = os.path.join(video_dir_path, video)
        print(f"Processing {video_path}")
        cap = cv2.VideoCapture(video_path)
        model_results = []
        rects = []
        while True:
            ret, frame = cap.read()
            if not ret: break
            rect = face_detect(frame)
            if rect is None: 
                model_results.append(None)
                rects.append(None)
                continue
            model_result = mask_detect(analyzer, frame, rect)
            model_results.append(model_result)
            rects.append(rect)
        cap.release()
        model_result_list[video] = model_results
        rect_list[video] = rects
    
    pickle.dump(model_result_list, open("/mask_test/model_result_list.pkl", "wb"))
    pickle.dump(rect_list, open("/mask_test/rect_list.pkl", "wb"))
        
def video_process(video_dir_path, q_size):
    model_result_list = pickle.load(open("/mask_test/model_result_list.pkl", "rb")) 
    rect_list = pickle.load(open("/mask_test/rect_list.pkl", "rb"))
    for video in os.listdir(video_dir_path):
        model_results = model_result_list[video]
        rects = rect_list[video]
        algorithm = StabilizationAlgorithm(q_size=q_size)
        video_path = os.path.join(video_dir_path, video)
        print(f"Processing {video_path}")
        cap = cv2.VideoCapture(video_path)
        combined_image_list = []
        frame_count = 0
        while True:
            ret, frame = cap.read()
            if not ret: break
            model_result = model_results[frame_count]
            rect = rects[frame_count]
            algo_result = algorithm.get_algorithm_result(model_result)
            
            combined_image = image_combination(frame, rect, model_result, algo_result)
            combined_image_list.append(combined_image)
            frame_count += 1
        
        video_name = video_path.split('/')[-1].split('.')[0]
        video_generate(video_name, combined_image_list)
        cap.release()

if __name__ == "__main__":
    # cal_model_result("test/Mask/mask")
    # queue_test()
    video_process("test/Mask/mask", 14)